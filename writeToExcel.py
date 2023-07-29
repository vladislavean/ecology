from openpyxl import load_workbook


class WriteToExcel:
    def __init__(self, data: dict, name_file: str, name_state: str):
        self.data = data
        self.name_state = name_state
        self.name_file = name_file

    def write(self):
        wb = load_workbook(filename=self.name_file)
        ws = wb.create_sheet(self.name_state) #имя листа = имя штата
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        # ws = wb[list_name]
        ws['A1'] = 'Штат: ' + self.name_state
        ws['A2'] = 'Годы:'
        ws['B2'] = 'Среднегодовая температура:'
        ws['C2'] = 'Отклонение от медианы:'
        row = 3
        for year, values in self.data[self.name_state][0].items():
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=values[0])
            ws.cell(row=row, column=3, value=values[1])
            row += 1
        ws.cell(row=row, column=1, value='Медиана:')
        ws.cell(row=row, column=2, value=self.data[self.name_state][1])
        ws.cell(row=row, column=3, value=self.data[self.name_state][2])
        wb.save(self.name_file)
        wb.close()